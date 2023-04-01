import json
import math
from pyclickup import ClickUp
import requests
from openpyxl import Workbook, load_workbook


API_TOKEN = "pk_36339500_YG104WEYKCZSMYYHMH635RWSW5780N37"

DAY_HOURS = 6



# def write_excel(lines):
#     wb = Workbook()

#     # grab the active worksheet
#     ws = wb.active

#     task_index = 1

#     for l in lines:
#         all_hours = 0
#         for t in lines[l]:
#             ws.cell(row=task_index,column=1).value = t["name"]
#             start_date = math.floor (all_hours / DAY_HOURS) + 1
#             column_count = 

#     ws.cell()
#     # Save the file
#     wb.save("sample.xlsx")


def tasks_line_by_asigne(tasks):
    tasks_lines = [t["name"] for t in tasks]
    

    with open("tasks.txt","w",encoding = "utf-8") as tasks_file:
        tasks_file.write("\n".join(tasks_lines))


def get_tasks(list_id):
    url = "https://api.clickup.com/api/v2/list/" + list_id + "/task"

    query = {
    "page": 0,
"order_by":"due_date",
"reverse":True
    }

    headers = {
    "Content-Type": "application/json",
    "Authorization": API_TOKEN
    }

    response = requests.get(url, headers=headers, params=query)

    data = response.json()
    tasks_line_by_asigne(data["tasks"])


def get_fields(id):
    url = "https://api.clickup.com/api/v2/list/" + id + "/field"



    headers = {
    "Content-Type": "application/json",
    "Authorization": API_TOKEN
    }

    response = requests.get(url, headers=headers)

    print_fields(response.json())



def write_from_excel(list_id):
    wb = load_workbook("")

    # grab the active worksheet
    ws = wb.active

    index = 2
    while True:
        if not ws.cell(row=index,column=1).value:
            break
    ws.cell()
    # Save the file
    wb.save("sample.xlsx")


def print_fields(fields):
    for f in fields["fields"]:
        print(f["name"],f["id"])

list_id = "900701270607"

#https://app.clickup.com/9007102490/v/li/900701270607


# https://app.clickup.com/24338528/v/li/900701187747
# https://app.clickup.com/9007102490/v/li/900701250251
get_fields(list_id)


